import os.path
import configparser
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
import openpyxl
from openpyxl import Workbook
from shopee_module import ShopeeAutomation

# If modifying these scopes, delete the file token.json.
SCOPES = ["https://www.googleapis.com/auth/drive"]
CHECKPOINT_FILE = 'processed_orders.json'

def validate_order_number(order_number):
    """
    Validate Shopee order number format.
    Shopee format: YYMMDD followed by alphanumeric (e.g., 2504226A23B55PX)
    
    Args:
        order_number: Order number string to validate
    
    Returns:
        bool: True if valid format, False otherwise
    """
    # Pattern: 6 digits (YYMMDD) + alphanumeric characters
    pattern = r'^\d{6}[A-Z0-9]+$'
    return bool(re.match(pattern, order_number.upper()))

def get_batch_orders():
    """
    Get order numbers from user with multiple input methods:
    1. Comma-separated list
    2. Newline-separated (traditional)
    3. From text file
    
    Returns:
        list: List of order numbers
    """
    print("\n" + "="*70)
    print("INPUT NOMOR PESANAN")
    print("="*70)
    print("\nPilih metode input:")
    print("1. Paste comma-separated (2504226A23B55PX, 2504226A34BUBPFX, ...)")
    print("2. Input satu per satu (tekan Enter dua kali untuk selesai)")
    print("3. Import dari file txt (orders.txt)")
    
    choice = input("\nPilih metode (1/2/3) [default: 2]: ").strip() or "2"
    
    order_numbers = []
    
    if choice == "1":
        # Comma-separated input
        print("\nPaste nomor pesanan (pisahkan dengan koma):")
        orders_input = input().strip()
        # Split by comma and clean up whitespace
        order_numbers = [order.strip() for order in orders_input.split(',') if order.strip()]
        
    elif choice == "3":
        # Read from file
        file_path = input("\nNama file [default: orders.txt]: ").strip() or "orders.txt"
        try:
            with open(file_path, 'r') as f:
                order_numbers = [line.strip() for line in f if line.strip()]
            print(f"‚úì Berhasil membaca {len(order_numbers)} pesanan dari {file_path}")
        except FileNotFoundError:
            print(f"‚úó File '{file_path}' tidak ditemukan")
            return []
        except Exception as e:
            print(f"‚úó Error membaca file: {e}")
            return []
    else:
        # Traditional one-by-one input (choice == "2")
        print("\nMasukkan nomor pesanan (satu per baris, Enter kosong untuk selesai):")
        while True:
            order = input("Nomor pesanan: ").strip()
            if not order:
                break
            order_numbers.append(order)
            print(f"  ‚úì Ditambahkan: {order}")
    
    # Validate order numbers
    if order_numbers:
        print(f"\nüìã Total {len(order_numbers)} pesanan ditemukan")
        print("\nüîç Validating order numbers...")
        
        valid_orders = []
        invalid_orders = []
        
        for order in order_numbers:
            if validate_order_number(order):
                valid_orders.append(order)
            else:
                invalid_orders.append(order)
        
        if invalid_orders:
            print(f"\n‚ö† WARNING: {len(invalid_orders)} pesanan dengan format mencurigakan:")
            for order in invalid_orders:
                print(f"  - {order}")
            
            confirm = input("\nLanjutkan dengan semua pesanan? (y/n) [default: y]: ").strip().lower() or 'y'
            if confirm != 'y':
                print("Hanya pesanan valid yang akan diproses.")
                return valid_orders
        
        print(f"‚úì Validasi selesai: {len(valid_orders)} valid, {len(invalid_orders)} warning")
        return order_numbers
    else:
        print("\n‚ö† Tidak ada nomor pesanan yang diinput")
        return []

def save_checkpoint(order_number, gdrive_link):
    """
    Save processed order to checkpoint file for resume capability.
    
    Args:
        order_number: Order number that was processed
        gdrive_link: Google Drive link for the screenshot
    """
    try:
        # Load existing checkpoint or create new
        if os.path.exists(CHECKPOINT_FILE):
            with open(CHECKPOINT_FILE, 'r') as f:
                checkpoint = json.load(f)
        else:
            checkpoint = {'processed_orders': [], 'timestamp': None}
        
        # Add new order
        checkpoint['processed_orders'].append({
            'order_number': order_number,
            'gdrive_link': gdrive_link,
            'timestamp': datetime.now().isoformat()
        })
        checkpoint['timestamp'] = datetime.now().isoformat()
        
        # Save checkpoint
        with open(CHECKPOINT_FILE, 'w') as f:
            json.dump(checkpoint, f, indent=2)
    except Exception as e:
        print(f"    ‚ö† Warning: Could not save checkpoint: {e}")

def load_checkpoint():
    """
    Load checkpoint to resume interrupted session.
    
    Returns:
        dict: Dictionary with 'processed_orders' list or empty dict
    """
    try:
        if os.path.exists(CHECKPOINT_FILE):
            with open(CHECKPOINT_FILE, 'r') as f:
                return json.load(f)
    except Exception as e:
        print(f"‚ö† Warning: Could not load checkpoint: {e}")
    return {'processed_orders': []}

def upload_to_gdrive_batch(service, file_paths, folder_id, max_workers=3):
    """
    Upload multiple files to Google Drive in parallel.
    
    Args:
        service: Google Drive API service object
        file_paths: List of file paths to upload
        folder_id: Google Drive folder ID
        max_workers: Maximum number of parallel uploads (default: 3)
    
    Returns:
        dict: Dictionary mapping file paths to their Google Drive links
    """
    results = {}
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all upload tasks
        future_to_path = {
            executor.submit(upload_to_gdrive, service, path, folder_id): path 
            for path in file_paths
        }
        
        # Collect results as they complete
        for future in as_completed(future_to_path):
            file_path = future_to_path[future]
            try:
                gdrive_link = future.result()
                results[file_path] = gdrive_link
            except Exception as e:
                print(f"    ‚úó Exception uploading {file_path}: {e}")
                results[file_path] = None
    
    return results

def check_duplicate_in_excel(order_number, excel_file='shopee_report.xlsx'):
    """
    Check if order number already exists in Excel report.
    
    Args:
        order_number: Order number to check
        excel_file: Path to Excel file
    
    Returns:
        bool: True if duplicate found, False otherwise
    """
    try:
        if not os.path.exists(excel_file):
            return False
        
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Check column B (OrderSN) starting from row 2
        for row in range(2, ws.max_row + 1):
            if ws[f'B{row}'].value == order_number:
                return True
        
        return False
    except Exception:
        return False

def get_gdrive_service():
    """
    Authenticates with the Google Drive API and returns a service object.
    """
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            # Make sure you have the credentials.json file from Google Cloud Console
            flow = InstalledAppFlow.from_client_secrets_file(
                "credentials.json", SCOPES, redirect_uri='urn:ietf:wg:oauth:2.0:oob'
            )
            # Use manual authorization flow
            auth_url, _ = flow.authorization_url(prompt='consent')
            
            print("\n" + "="*70)
            print("AUTHORIZATION REQUIRED")
            print("="*70)
            print("\nPlease visit this URL to authorize this application:")
            print("\n" + auth_url + "\n")
            print("After authorization, you will get a code.")
            code = input("Enter the authorization code here: ").strip()
            flow.fetch_token(code=code)
            creds = flow.credentials
        
        # Save the credentials for the next run
        with open("token.json", "w") as token:
            token.write(creds.to_json())

    try:
        service = build("drive", "v3", credentials=creds)
        print("Successfully connected to Google Drive API.")
        return service
    except HttpError as error:
        print(f"An error occurred: {error}")
        return None

def load_config():
    """Load configuration from config.ini file."""
    config = configparser.ConfigParser()
    config.read('config.ini')
    return config

def upload_to_gdrive(service, file_path, folder_id, max_retries=3):
    """
    Uploads a file to Google Drive with retry mechanism and returns the shareable link.
    
    Args:
        service: Google Drive API service object
        file_path: Path to the file to upload
        folder_id: ID of the Google Drive folder to upload to
        max_retries: Maximum number of retry attempts (default: 3)
    
    Returns:
        str: Shareable link to the uploaded file, or None if upload fails
    """
    file_name = os.path.basename(file_path)
    
    for attempt in range(1, max_retries + 1):
        try:
            file_metadata = {
                'name': file_name,
                'parents': [folder_id]
            }
            
            media = MediaFileUpload(file_path, resumable=True)
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id, webViewLink'
            ).execute()
            
            # Make the file accessible to anyone with the link
            permission = {
                'type': 'anyone',
                'role': 'reader'
            }
            service.permissions().create(
                fileId=file.get('id'),
                body=permission
            ).execute()
            
            print(f"    ‚úì Uploaded: {file_name}")
            return file.get('webViewLink')
            
        except HttpError as error:
            if attempt < max_retries:
                wait_time = 2 ** attempt  # Exponential backoff: 2, 4, 8 seconds
                print(f"    ‚ö† Upload attempt {attempt} failed, retrying in {wait_time}s...")
                time.sleep(wait_time)
            else:
                print(f"    ‚úó Upload failed after {max_retries} attempts: {error}")
                return None
    
    return None

def create_excel_report(order_data, output_file='shopee_report.xlsx'):
    """
    Creates an Excel report with order numbers and Google Drive links.
    Format follows Shopee CS template with 3 columns: No, OrderSN, Bukti
    If the file exists, it will append new data. Otherwise, create new file.
    
    Args:
        order_data: List of dictionaries with 'order_number' and 'gdrive_link'
        output_file: Name of the output Excel file
    
    Returns:
        str: Path to the created Excel file
    """
    # Check if file already exists
    if os.path.exists(output_file):
        print(f"File '{output_file}' sudah ada, akan menambahkan data baru...")
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Find the next empty row (skip header row)
        next_row = ws.max_row + 1
        
        # Get current number for sequential numbering
        current_no = next_row - 1  # Assuming row 1 is header
        
        # Add new data
        for data in order_data:
            ws[f'A{next_row}'] = current_no  # Sequential number
            ws[f'B{next_row}'] = data['order_number']
            ws[f'C{next_row}'] = data['gdrive_link']
            next_row += 1
            current_no += 1
        
        print(f"‚úì Menambahkan {len(order_data)} data baru ke file existing")
    else:
        print(f"Membuat file Excel baru: {output_file}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add headers (Row 1) - matching Shopee template
        ws['A1'] = "No"
        ws['B1'] = "OrderSN/ Nomor Pesanan"
        ws['C1'] = "Bukti pembeli sudah menerima pesanan\n- Screenshot yang menunjukkan pembeli sudah mengonfirmasi menerima produk non fisik. Screenshot harus dari Chat di Shopee, screenshot dari platform lain (cth Whatsapp) tidak akan diproses\n- Masukkan foto kedalam google drive dan salin ulang link kedalam kolom dibawah ini\n- Pastikan google drive tidak terkunci sehingga dapat diakses oleh Tim Shopee"
        
        # Make headers bold and wrap text
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = openpyxl.styles.Font(bold=True)
            ws[cell].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        
        # Add data starting from row 2
        for idx, data in enumerate(order_data, start=2):
            ws[f'A{idx}'] = idx - 1  # Sequential number starting from 1
            ws[f'B{idx}'] = data['order_number']
            ws[f'C{idx}'] = data['gdrive_link']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 100
        
        # Set row height for header
        ws.row_dimensions[1].height = 80
    
    wb.save(output_file)
    print(f"‚úì Excel report created/updated: {output_file}")
    return output_file

def main():
    """Main function to run the full automation workflow."""
    print("\n" + "="*70)
    print("üöÄ SHOPEE AUTOMATION - ENHANCED VERSION")
    print("="*70)
    
    # Load configuration
    config = load_config()
    folder_id = config.get('GOOGLE_DRIVE', 'FOLDER_ID')
    username = config.get('SHOPEE', 'USERNAME')
    password = config.get('SHOPEE', 'PASSWORD')
    chrome_profile = config.get('SHOPEE', 'CHROME_PROFILE', fallback='Default')
    
    # Check if credentials are configured
    if username == 'your_shopee_username' or password == 'your_shopee_password':
        print("\n‚úó ERROR: Shopee credentials not configured!")
        print("Please edit config.ini and add your Shopee username and password.")
        return
    
    # Check for resume capability
    checkpoint = load_checkpoint()
    if checkpoint.get('processed_orders'):
        print(f"\nüìå Found checkpoint with {len(checkpoint['processed_orders'])} processed orders")
        resume = input("Resume from checkpoint? (y/n) [default: n]: ").strip().lower()
        if resume == 'y':
            processed_order_numbers = [o['order_number'] for o in checkpoint['processed_orders']]
            print(f"‚úì Will skip {len(processed_order_numbers)} already processed orders")
        else:
            processed_order_numbers = []
            # Clear checkpoint
            if os.path.exists(CHECKPOINT_FILE):
                os.remove(CHECKPOINT_FILE)
    else:
        processed_order_numbers = []
    
    # Step 1: Connect to Google Drive
    print("\n[1/5] üì° Connecting to Google Drive...")
    gdrive_service = get_gdrive_service()
    if not gdrive_service:
        print("‚úó Could not connect to Google Drive. Aborting.")
        return
    print("‚úì Google Drive connected!")
    
    # Step 2: Initialize Shopee automation
    print("\n[2/5] üåê Initializing Shopee automation...")
    shopee = ShopeeAutomation(username, password, headless=False, chrome_profile=chrome_profile)
    shopee.start_browser()
    
    try:
        # Step 3: Login to Shopee
        print("\n[3/5] üîê Logging in to Shopee Seller Centre...")
        if not shopee.login():
            print("‚úó Login failed. Aborting.")
            return
        
        # Step 4: Get orders (auto-detect or batch input)
        print("\n[4/5] üì¶ Getting orders...")
        
        # Check auto-detect setting
        auto_detect = config.getboolean('AUTOMATION', 'AUTO_DETECT_ORDERS', fallback=True)
        
        # First try auto-detect from Shopee page if enabled
        if auto_detect:
            print("‚Ñπ Auto-detect enabled (can be disabled in config.ini)")
            order_numbers = shopee.get_orders_to_ship(auto_detect=True)
        else:
            print("‚Ñπ Auto-detect disabled, using manual/batch input")
            order_numbers = []
        
        # If auto-detect returned nothing, use batch input
        if not order_numbers:
            order_numbers = get_batch_orders()
        
        if not order_numbers:
            print("No orders to process. Exiting.")
            return
        
        # Filter out already processed orders
        if processed_order_numbers:
            original_count = len(order_numbers)
            order_numbers = [o for o in order_numbers if o not in processed_order_numbers]
            skipped = original_count - len(order_numbers)
            if skipped > 0:
                print(f"‚è≠ Skipping {skipped} already processed orders")
        
        if not order_numbers:
            print("All orders already processed!")
            return
        
        # Check for duplicates in Excel
        print("\nüîç Checking for duplicates in existing Excel...")
        duplicates = []
        for order in order_numbers:
            if check_duplicate_in_excel(order):
                duplicates.append(order)
        
        if duplicates:
            print(f"\n‚ö† WARNING: {len(duplicates)} order(s) already in Excel report:")
            for dup in duplicates[:5]:  # Show first 5
                print(f"  - {dup}")
            if len(duplicates) > 5:
                print(f"  ... and {len(duplicates) - 5} more")
            
            confirm = input("\nProcess anyway? (y/n) [default: n]: ").strip().lower()
            if confirm != 'y':
                # Remove duplicates
                order_numbers = [o for o in order_numbers if o not in duplicates]
                print(f"‚úì Removed {len(duplicates)} duplicates, {len(order_numbers)} orders remaining")
                
                if not order_numbers:
                    print("No orders to process. Exiting.")
                    return
        
        # Process orders with progress tracking
        order_data = []
        failed_orders = []
        screenshots_folder = 'screenshots'
        total_orders = len(order_numbers)
        start_time = time.time()
        
        print(f"\n{'='*70}")
        print(f"üì∏ PROCESSING {total_orders} ORDERS")
        print(f"{'='*70}\n")
        
        for i, order_number in enumerate(order_numbers, 1):
            # Progress indicator
            progress_pct = (i / total_orders) * 100
            elapsed = time.time() - start_time
            if i > 1:
                avg_time_per_order = elapsed / (i - 1)
                remaining_orders = total_orders - i
                eta_seconds = avg_time_per_order * remaining_orders
                eta_minutes = int(eta_seconds // 60)
                eta_seconds_remainder = int(eta_seconds % 60)
                eta_str = f"ETA: {eta_minutes}m {eta_seconds_remainder}s"
            else:
                eta_str = "ETA: calculating..."
            
            print(f"\n{'‚îÄ'*70}")
            print(f"[{i}/{total_orders}] ({progress_pct:.0f}%) | {eta_str}")
            print(f"Order: {order_number}")
            print(f"{'‚îÄ'*70}")
            
            # Take screenshot
            screenshot_path = shopee.take_chat_screenshot(order_number, screenshots_folder)
            
            if screenshot_path:
                # Upload to Google Drive with retry
                print(f"  üì§ Uploading to Google Drive...")
                gdrive_link = upload_to_gdrive(gdrive_service, screenshot_path, folder_id)
                
                if gdrive_link:
                    order_data.append({
                        'order_number': order_number,
                        'gdrive_link': gdrive_link
                    })
                    # Save checkpoint
                    save_checkpoint(order_number, gdrive_link)
                    print(f"  ‚úÖ Order {order_number} processed successfully!")
                else:
                    failed_orders.append({'order': order_number, 'reason': 'Upload failed'})
                    print(f"  ‚ùå Failed to upload screenshot")
            else:
                failed_orders.append({'order': order_number, 'reason': 'Screenshot failed'})
                print(f"  ‚ùå Failed to take screenshot")
        
        # Step 5: Generate Excel report
        print(f"\n{'='*70}")
        print("[5/5] üìä Generating Excel report...")
        print(f"{'='*70}")
        
        if order_data:
            excel_file = create_excel_report(order_data, 'shopee_report.xlsx')
            
            # Final summary
            total_time = time.time() - start_time
            minutes = int(total_time // 60)
            seconds = int(total_time % 60)
            
            print(f"\n{'='*70}")
            print("‚úÖ AUTOMATION COMPLETED!")
            print(f"{'='*70}")
            print(f"\nüìä SUMMARY:")
            print(f"  ‚úì Total processed: {len(order_data)}/{total_orders}")
            print(f"  ‚úì Successful: {len(order_data)}")
            print(f"  ‚úó Failed: {len(failed_orders)}")
            print(f"  ‚è± Total time: {minutes}m {seconds}s")
            print(f"  üìÅ Excel report: {excel_file}")
            
            if failed_orders:
                print(f"\n‚ùå FAILED ORDERS:")
                for fail in failed_orders:
                    print(f"  - {fail['order']}: {fail['reason']}")
                
                # Save failed orders to file
                with open('failed_orders.txt', 'w') as f:
                    f.write(f"Failed orders ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})\n")
                    f.write("="*50 + "\n\n")
                    for fail in failed_orders:
                        f.write(f"{fail['order']} - {fail['reason']}\n")
                print(f"\n  üìù Failed orders saved to: failed_orders.txt")
            
            print(f"\nüìã NEXT STEPS:")
            print(f"  1. Open {excel_file}")
            print(f"  2. Verify all data is correct")
            print(f"  3. Submit the report to Shopee CS")
            
            if failed_orders:
                print(f"  4. Retry failed orders from failed_orders.txt")
        else:
            print("\n‚ö† No orders were successfully processed.")
            
    except KeyboardInterrupt:
        print("\n\n‚ö† Process interrupted by user.")
        print("üíæ Progress has been saved. You can resume later.")
    except Exception as e:
        print(f"\n‚úó Error during automation: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # Cleanup
        print("\nClosing browser...")
        shopee.close_browser()
        print("‚úì Automation finished.")


if __name__ == "__main__":
    main()

